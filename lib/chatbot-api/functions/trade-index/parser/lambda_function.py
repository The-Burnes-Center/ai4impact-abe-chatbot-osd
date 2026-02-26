"""
Trade Index Parser Lambda: triggered on S3 upload of .xlsx to trade-index/ prefix.
Schema-flexible — accepts any headers. Writes rows + META to DynamoDB.
"""
import io
import json
import os
from datetime import datetime, timezone

import boto3
from openpyxl import load_workbook

from models import excel_column_to_field, row_dict_from_excel_row

S3 = boto3.client("s3")
DDB = boto3.resource("dynamodb")
BUCKET = os.environ["BUCKET"]
TABLE_NAME = os.environ["TABLE_NAME"]
PK = "INDEX"
SK_META = "META"
BATCH_SIZE = 25


def _clear_table(table) -> None:
    """Delete all row items with pk=INDEX; keep META so status shows PROCESSING."""
    keys_to_delete = []
    paginator = boto3.client("dynamodb").get_paginator("query")
    for page in paginator.paginate(
        TableName=TABLE_NAME,
        KeyConditionExpression="pk = :pk",
        ExpressionAttributeValues={":pk": {"S": PK}},
        ProjectionExpression="pk, sk",
    ):
        for item in page.get("Items", []):
            if item.get("sk", {}).get("S") == SK_META:
                continue
            keys_to_delete.append({"pk": item["pk"], "sk": item["sk"]})
    for i in range(0, len(keys_to_delete), BATCH_SIZE):
        chunk = keys_to_delete[i : i + BATCH_SIZE]
        request_items = {TABLE_NAME: [{"DeleteRequest": {"Key": k}} for k in chunk]}
        boto3.client("dynamodb").batch_write_item(RequestItems=request_items)


def _put_meta(table, row_count: int, last_updated: str, error: str | None = None, status: str | None = None) -> None:
    item: dict = {"pk": PK, "sk": SK_META, "row_count": row_count, "last_updated": last_updated}
    if status is not None:
        item["status"] = status
    if error is not None:
        item["error"] = error
    table.put_item(Item=item)


def _serialize_value(v) -> str:
    if v is None:
        return ""
    return str(v)


def lambda_handler(event, context):
    """Process S3 PutObject for .xlsx in trade-index/ prefix; write to DynamoDB."""
    table = DDB.Table(TABLE_NAME)
    for record in event.get("Records", []):
        bucket = record["s3"]["bucket"]["name"]
        key = record["s3"]["object"]["key"]
        if not key.lower().endswith(".xlsx"):
            continue

        try:
            obj = S3.get_object(Bucket=bucket, Key=key)
            body = obj["Body"].read()
            wb = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            headers = [str(h).strip() if h is not None else "" for h in headers]

            non_empty = [h for h in headers if h]
            if len(non_empty) < 2:
                err = f"Trade Index file has only {len(non_empty)} header column(s); expected at least 2."
                _put_meta(table, 0, datetime.now(timezone.utc).isoformat(), error=err, status="ERROR")
                wb.close()
                return {"statusCode": 200, "body": json.dumps({"status": "error", "message": err})}

            now_start = datetime.now(timezone.utc).isoformat()
            _put_meta(table, 0, now_start, error=None, status="PROCESSING")

            col_names = [excel_column_to_field(h) for h in non_empty]
            rows_out = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = row_dict_from_excel_row(headers, row)
                if not row_dict or all(v == "" for v in row_dict.values()):
                    continue
                rows_out.append(row_dict)

            wb.close()
            _clear_table(table)

            now = datetime.now(timezone.utc).isoformat()
            _put_meta(table, len(rows_out), now, error=None, status="COMPLETE")

            # Store column names in META for preview
            table.update_item(
                Key={"pk": PK, "sk": SK_META},
                UpdateExpression="SET columns = :c",
                ExpressionAttributeValues={":c": col_names},
            )

            for offset in range(0, len(rows_out), BATCH_SIZE):
                chunk = rows_out[offset : offset + BATCH_SIZE]
                with table.batch_writer() as writer:
                    for j, row in enumerate(chunk):
                        item = {"pk": PK, "sk": str(offset + j)}
                        for k, v in row.items():
                            item[k] = _serialize_value(v)
                        writer.put_item(Item=item)

            return {"statusCode": 200, "body": json.dumps({"status": "ok", "row_count": len(rows_out)})}

        except Exception as e:
            print(f"Trade Index parser error: {e}")
            try:
                _put_meta(table, 0, datetime.now(timezone.utc).isoformat(), error=str(e), status="ERROR")
            except Exception as meta_err:
                print(f"Failed to write meta error: {meta_err}")
            return {"statusCode": 200, "body": json.dumps({"status": "error", "message": str(e)})}

    return {"statusCode": 200, "body": "{}"}

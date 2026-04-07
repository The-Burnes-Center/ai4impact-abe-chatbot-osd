"""
Excel Index Parser Lambda -- S3 event-driven ingestion of .xlsx files into DynamoDB.

Triggered automatically by S3 notifications when a file is created or deleted
under the ``indexes/`` prefix. The S3 key must follow the convention
``indexes/{index_id}/latest.xlsx``; other paths are silently ignored.

Processing pipeline:
  1. Extract ``index_id`` from the S3 key.
  2. For delete events: clear all data rows and remove the tool-registry entry.
  3. For create/update events:
     a. Download and parse the Excel file with openpyxl.
     b. Validate that at least 2 header columns exist.
     c. Write a PROCESSING status to the META item.
     d. Clear stale rows from any previous version of this index.
     e. Batch-write all new rows to DynamoDB.
     f. Update the META item with COMPLETE status, row count, and column list.
     g. Register the index in the tool registry so the chat agent can discover
        and query it. The registry call triggers AI-generated descriptions of
        the index contents based on column names and sample rows.

DynamoDB layout (shared table, partitioned by index_id):
  - ``pk=index_id, sk=META`` -- status, row_count, last_updated, column list
  - ``pk=index_id, sk=0..N`` -- one item per Excel row

The META item is intentionally preserved during ``_clear_index`` so that the
UI can display status/error information even while the index is being
reprocessed. It is overwritten (not deleted) once the new parse completes.
"""
import io
import json
import os
import re
from datetime import datetime, timezone

import boto3
from openpyxl import load_workbook

from models import excel_column_to_field, row_dict_from_excel_row
from tool_registry import write_to_registry, delete_from_registry

S3 = boto3.client("s3")
DDB = boto3.resource("dynamodb")
BUCKET = os.environ["BUCKET"]
TABLE_NAME = os.environ["TABLE_NAME"]
SK_META = "META"
BATCH_SIZE = 25

_INDEX_ID_RE = re.compile(r"^indexes/([^/]+)/")


def _extract_index_id(key: str) -> str | None:
    """Extract index_id from S3 key like indexes/{index_id}/latest.xlsx."""
    m = _INDEX_ID_RE.match(key)
    return m.group(1) if m else None


def _index_id_to_display_name(index_id: str) -> str:
    """Convert snake_case index_id to Title Case display name."""
    return index_id.replace("_", " ").title()


def _clear_index(table, index_id: str) -> None:
    """Delete all data-row items for the given index_id, preserving the META item.

    The META item (sk=META) is kept so the UI can still display index status
    and error information during reprocessing. It will be overwritten by
    ``_put_meta`` once parsing completes or fails.
    """
    keys_to_delete = []
    paginator = boto3.client("dynamodb").get_paginator("query")
    for page in paginator.paginate(
        TableName=TABLE_NAME,
        KeyConditionExpression="pk = :pk",
        ExpressionAttributeValues={":pk": {"S": index_id}},
        ProjectionExpression="pk, sk",
    ):
        for item in page.get("Items", []):
            if item.get("sk", {}).get("S") == SK_META:
                continue
            keys_to_delete.append({"pk": item["pk"], "sk": item["sk"]})
    for i in range(0, len(keys_to_delete), BATCH_SIZE):
        chunk = keys_to_delete[i : i + BATCH_SIZE]
        request_items = {TABLE_NAME: [{"DeleteRequest": {"Key": k}} for k in chunk]}
        boto3.client("dynamodb").batch_write_item(RequestItems=request_items)


def _put_meta(table, index_id: str, row_count: int, last_updated: str,
              error: str | None = None, status: str | None = None) -> None:
    """Write or overwrite the META item for an index with current status info."""
    item: dict = {"pk": index_id, "sk": SK_META, "row_count": row_count, "last_updated": last_updated}
    if status is not None:
        item["status"] = status
    if error is not None:
        item["error"] = error
    table.put_item(Item=item)


def _serialize_value(v) -> str:
    """Convert any cell value to a string for DynamoDB storage (None becomes empty string)."""
    if v is None:
        return ""
    return str(v)


def lambda_handler(event, context):
    """Process S3 event records for Excel index files.

    Handles both ObjectCreated and ObjectRemoved events. For each record:
      - ObjectRemoved: clears the index data and deregisters the tool.
      - ObjectCreated: downloads the .xlsx, parses it, writes rows to DynamoDB,
        and registers the index in the tool registry (which triggers AI
        description generation from column names and sample rows).

    Returns a 200 response with status details for the last processed record.
    """
    table = DDB.Table(TABLE_NAME)
    for record in event.get("Records", []):
        event_name = record.get("eventName", "")
        bucket = record["s3"]["bucket"]["name"]
        key = record["s3"]["object"]["key"]
        if not key.lower().endswith(".xlsx"):
            continue

        index_id = _extract_index_id(key)
        if not index_id:
            print(f"Could not extract index_id from key: {key}")
            continue

        display_name = _index_id_to_display_name(index_id)

        if "ObjectRemoved" in event_name:
            print(f"Delete event for {key}; clearing index '{index_id}' and registry.")
            _clear_index(table, index_id)
            _put_meta(table, index_id, 0, datetime.now(timezone.utc).isoformat(), status="NO_DATA")
            delete_from_registry(index_id)
            return {"statusCode": 200, "body": json.dumps({"status": "deleted", "index_id": index_id})}

        try:
            obj = S3.get_object(Bucket=bucket, Key=key)
            body = obj["Body"].read()
            wb = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            headers = [str(h).strip() if h is not None else "" for h in headers]

            non_empty = [h for h in headers if h]
            if len(non_empty) < 2:
                err = f"Index '{index_id}' file has only {len(non_empty)} header column(s); expected at least 2."
                _put_meta(table, index_id, 0, datetime.now(timezone.utc).isoformat(), error=err, status="ERROR")
                wb.close()
                return {"statusCode": 200, "body": json.dumps({"status": "error", "message": err})}

            now_start = datetime.now(timezone.utc).isoformat()
            _put_meta(table, index_id, 0, now_start, error=None, status="PROCESSING")

            col_names = [excel_column_to_field(h) for h in non_empty]
            rows_out = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = row_dict_from_excel_row(headers, row)
                if not row_dict or all(v == "" for v in row_dict.values()):
                    continue
                rows_out.append(row_dict)

            wb.close()
            _clear_index(table, index_id)

            now = datetime.now(timezone.utc).isoformat()
            _put_meta(table, index_id, len(rows_out), now, error=None, status="COMPLETE")

            table.update_item(
                Key={"pk": index_id, "sk": SK_META},
                UpdateExpression="SET #col = :c",
                ExpressionAttributeNames={"#col": "columns"},
                ExpressionAttributeValues={":c": col_names},
            )

            for offset in range(0, len(rows_out), BATCH_SIZE):
                chunk = rows_out[offset : offset + BATCH_SIZE]
                with table.batch_writer() as writer:
                    for j, row in enumerate(chunk):
                        item = {"pk": index_id, "sk": str(offset + j)}
                        for k, v in row.items():
                            item[k] = _serialize_value(v)
                        writer.put_item(Item=item)

            write_to_registry(index_id, display_name, col_names, len(rows_out), sample_rows=rows_out[:5])
            print(f"Parsed index '{index_id}': {len(rows_out)} rows, {len(col_names)} columns.")
            return {"statusCode": 200, "body": json.dumps({"status": "ok", "index_id": index_id, "row_count": len(rows_out)})}

        except Exception as e:
            print(f"Parser error for index '{index_id}': {e}")
            try:
                _put_meta(table, index_id, 0, datetime.now(timezone.utc).isoformat(), error=str(e), status="ERROR")
            except Exception as meta_err:
                print(f"Failed to write meta error: {meta_err}")
            return {"statusCode": 200, "body": json.dumps({"status": "error", "message": str(e)})}

    return {"statusCode": 200, "body": "{}"}

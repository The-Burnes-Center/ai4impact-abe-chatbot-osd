"""
Parser Lambda: triggered on S3 upload of .xlsx to contract index bucket.
Validates schema with Pydantic, writes rows + meta to DynamoDB.
"""
import io
import json
import os
from datetime import datetime, timezone

import boto3
from openpyxl import load_workbook

from models import (
    SWC_INDEX_COLUMNS,
    SWCRow,
    excel_column_to_field,
    row_dict_from_excel_row,
)

S3 = boto3.client("s3")
DDB = boto3.resource("dynamodb")
BUCKET = os.environ["BUCKET"]
TABLE_NAME = os.environ["TABLE_NAME"]
ERROR_KEY = "swc-index/error.json"
PK = "INDEX"
SK_META = "META"
BATCH_SIZE = 25


def _validate_headers(headers: list[str]) -> tuple[bool, str]:
    """Check that Excel headers match expected SWC Index columns."""
    normalized = [excel_column_to_field(h) for h in headers if h]
    expected_set = set(excel_column_to_field(c) for c in SWC_INDEX_COLUMNS)
    actual_set = set(normalized)
    if expected_set - actual_set:
        missing = expected_set - actual_set
        return False, f"Missing columns: {missing}"
    return True, ""


def _clear_table(table) -> None:
    """Delete all items with pk=INDEX (rows + META)."""
    pk_name = "pk"
    sk_name = "sk"
    keys_to_delete = []
    paginator = boto3.client("dynamodb").get_paginator("query")
    for page in paginator.paginate(
        TableName=TABLE_NAME,
        KeyConditionExpression="pk = :pk",
        ExpressionAttributeValues={":pk": {"S": PK}},
        ProjectionExpression="pk, sk",
    ):
        for item in page.get("Items", []):
            keys_to_delete.append(
                {pk_name: item[pk_name], sk_name: item[sk_name]}
            )
    for i in range(0, len(keys_to_delete), BATCH_SIZE):
        chunk = keys_to_delete[i : i + BATCH_SIZE]
        request_items = {
            TABLE_NAME: [
                {"DeleteRequest": {"Key": k}}
                for k in chunk
            ]
        }
        boto3.client("dynamodb").batch_write_item(RequestItems=request_items)


def _put_meta(table, row_count: int, last_updated: str, error: str | None = None) -> None:
    """Write META item."""
    item = {
        "pk": PK,
        "sk": SK_META,
        "row_count": row_count,
        "last_updated": last_updated,
    }
    if error is not None:
        item["error"] = error
    table.put_item(Item=item)


def _serialize_value(v) -> str:
    """DynamoDB string attribute."""
    if v is None:
        return ""
    return str(v)


def lambda_handler(event, context):
    """Process S3 PutObject for .xlsx in contract index bucket; write to DynamoDB."""
    table = DDB.Table(TABLE_NAME)
    for record in event.get("Records", []):
        bucket = record["s3"]["bucket"]["name"]
        key = record["s3"]["object"]["key"]
        if not key.lower().endswith(".xlsx"):
            continue

        try:
            obj = S3.get_object(Bucket=bucket, Key=key)
            body = obj["Body"].read()
            wb = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            headers = [str(h).strip() if h is not None else "" for h in headers]

            ok, err = _validate_headers(headers)
            if not ok:
                _put_meta(table, 0, datetime.now(timezone.utc).isoformat(), error=err)
                wb.close()
                return {"statusCode": 200, "body": json.dumps({"status": "error", "message": err})}

            rows_out = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = row_dict_from_excel_row(headers, row)
                if not row_dict:
                    continue
                try:
                    validated = SWCRow.model_validate(row_dict)
                    rows_out.append(validated.model_dump())
                except Exception as e:
                    print(f"Row validation skipped: {e}")
                    continue

            wb.close()

            _clear_table(table)

            now = datetime.now(timezone.utc).isoformat()
            _put_meta(table, len(rows_out), now, error=None)

            for offset in range(0, len(rows_out), BATCH_SIZE):
                chunk = rows_out[offset : offset + BATCH_SIZE]
                with table.batch_writer() as writer:
                    for j, row in enumerate(chunk):
                        item = {"pk": PK, "sk": str(offset + j)}
                        for k, v in row.items():
                            item[k] = _serialize_value(v)
                        writer.put_item(Item=item)

            return {"statusCode": 200, "body": json.dumps({"status": "ok", "row_count": len(rows_out)})}

        except Exception as e:
            print(f"Parser error: {e}")
            try:
                _put_meta(table, 0, datetime.now(timezone.utc).isoformat(), error=str(e))
            except Exception as meta_err:
                print(f"Failed to write meta error: {meta_err}")
            return {"statusCode": 200, "body": json.dumps({"status": "error", "message": str(e)})}

    return {"statusCode": 200, "body": "{}"}
